#!/usr/bin/env python3
import argparse
import openpyxl
from openpyxl.utils import column_index_from_string
import os
import glob
import sys

def parse_range(val, is_column=False):
    """
    Parses strings like "3", "3..", "..5", "3..5", "A", "A..", "A..C".
    Returns (min_val, max_val) as integers or None (for open ends).
    """
    if not val:
        return None, None

    if '..' in val:
        parts = val.split('..')
        start = parts[0]
        end = parts[1]
    else:
        start = val
        end = val

    # Helper to convert col letter to index or keep row number
    def convert(v):
        if not v: return None
        if is_column:
            try:
                return column_index_from_string(v.upper())
            except ValueError:
                print(f"Error: Invalid column letter '{v}'")
                sys.exit(1)
        else:
            try:
                return int(v)
            except ValueError:
                print(f"Error: Invalid row number '{v}'")
                sys.exit(1)

    return convert(start), convert(end)

def extract_data(files, col_arg, row_arg):
    # Expand wildcards for Windows
    expanded_files = []
    for f in files:
        if '*' in f or '?' in f:
            expanded_files.extend(glob.glob(f))
        else:
            expanded_files.append(f)

    if not expanded_files:
        print("No files found matching the pattern.")
        return

    # Parse ranges once
    min_c, max_c = parse_range(col_arg, is_column=True)
    min_r, max_r = parse_range(row_arg, is_column=False)

    for filepath in expanded_files:
        if not os.path.exists(filepath):
            continue

        try:
            print(f"Processing: {filepath}...")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active

            # Handle open-ended ranges (e.g., "3.." becomes 3 to max_row)
            # If argument wasn't provided (None), default to full range (1 to max)
            final_min_row = min_r if min_r else 1
            final_max_row = max_r if max_r else sheet.max_row
            
            final_min_col = min_c if min_c else 1
            final_max_col = max_c if max_c else sheet.max_column

            extracted_rows = []

            # iter_rows handles the slicing efficiently
            row_iterator = sheet.iter_rows(
                min_row=final_min_row,
                max_row=final_max_row,
                min_col=final_min_col,
                max_col=final_max_col,
                values_only=True
            )

            for row in row_iterator:
                # Filter out None values and convert to string
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                # Join columns with a Tab or Comma (using Tab here for clarity)
                extracted_rows.append("\t".join(cleaned_row))

            # Generate Filename
            base_name = os.path.splitext(os.path.basename(filepath))[0]
            
            # Construct a descriptive suffix based on arguments
            c_tag = f"_Col{col_arg}" if col_arg else ""
            r_tag = f"_Row{row_arg}" if row_arg else ""
            
            # Sanitize filename (replace .. with -to-)
            safe_c = c_tag.replace("..", "-")
            safe_r = r_tag.replace("..", "-")
            
            out_filename = f"{base_name}{safe_c}{safe_r}.txt"

            with open(out_filename, 'w', encoding='utf-8') as f:
                f.write('\n'.join(extracted_rows))
            
            print(f"  -> Saved to: {out_filename}")

        except Exception as e:
            print(f"Error processing {filepath}: {e}")

def main():
    parser = argparse.ArgumentParser(
        description="Extract specific Rows/Columns or Ranges from XLSX. Syntax: A, A..C, 3, 3..10"
    )
    
    parser.add_argument('files', nargs='+', help='XLSX files (e.g., *.xlsx)')
    parser.add_argument('-c', '--column', help='Column(s). Ex: A, A..C, C.., ..F')
    parser.add_argument('-r', '--row', help='Row(s). Ex: 1, 1..5, 5.., ..10')

    args = parser.parse_args()

    if not args.column and not args.row:
        print("Error: You must provide at least -c or -r (or both).")
        sys.exit(1)

    extract_data(args.files, args.column, args.row)

if __name__ == "__main__":
    main()